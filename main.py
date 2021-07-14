from datetime import *
from calendar import *
from openpyxl import *


wb = Workbook()
ws = wb.active
today = date.today()
number_of_days = monthrange(today.year, today.month)[1]
#number_of_days = 7
week_count = 6
row = 1
week_set = {}
sheet_header = ["Date", "Day", "Morning Shift", "Morning Shift", "Second Shift", "Second Shift"]


for head_date in range(1, len(sheet_header)+1):
    ws.cell(1, head_date).value = sheet_header[head_date-1]


def weekly_data(row_num, week_date, week_number):
    row_num = row_num+1
    date_data = f"{week_date}-{today.month}-{today.year}"
    week_day_name = datetime.strptime(date_data, '%d-%m-%Y').weekday()
    print(day_name[week_day_name])
    print(date_data)
    ws.cell(row_num, 1).value = date_data
    ws.cell(row_num, 2).value = (day_name[week_day_name])
    morning_person_data = week_set[f"Week{week_number}_M"]
    x_morning_person_data = morning_person_data.split(",")
    ws.cell(row_num, 3).value = x_morning_person_data[0]
    ws.cell(row_num, 4).value = x_morning_person_data[1]
    late_person_data = week_set[f"Week{week_number}_S"]
    x_late_person_data = late_person_data.split(",")
    ws.cell(row_num, 5).value = x_late_person_data[0]
    ws.cell(row_num, 6).value = x_late_person_data[1]


for i in range(1, week_count):
    actual_M_value = input(f"Enter Week - {i} Morning shift Person details with , separated :")
    temp1 = actual_M_value.find(",")
    if temp1 > 0:
        week_set[f"Week{i}_M"] = actual_M_value
    else:
        print("Enter the valid input")
        break

    actual_L_value = input(f"Enter Week - {i} Late shift Person details with , separated :")
    temp2 = actual_L_value.find(",")
    if temp2 > 0:
        week_set[f"Week{i}_S"] = actual_L_value
    else:
        print("Enter the valid input")
        break
print(week_set)

for i in range(1, number_of_days + 1):

    if i <= 7:
        Week_Number = 1
        weekly_data(i, row, Week_Number)
        i+1
        row = row+1

    elif i <= 8 or i <= 14:
        Week_Number = 2
        weekly_data(i, row, Week_Number)
        i+1
        row = row + 1

    elif i <= 15 or i <= 21:
        Week_Number = 3
        weekly_data(i, row, Week_Number)
        i + 1
        row = row + 1
    elif i <= 21 or i <= 28:
        Week_Number = 4
        weekly_data(i, row, Week_Number)
        i + 1
        row = row + 1
    elif i <= 28 or i <= 31:
        Week_Number = 5
        weekly_data(i, row, Week_Number)
        i + 1
        row = row + 1
    else:
        print("Month over")

wb.save("/Users/mani/Desktop/Python_Project/On-call/Shift_Roaster.xlsx")
